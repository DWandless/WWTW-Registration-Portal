from pathlib import Path
import streamlit as st
import pandas as pd
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from db import get_supabase
import xml.etree.ElementTree as ET
import math
import os
import re
import unicodedata


# -------------------------------------------------------------------
# Paths + Static Assets
# -------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
ASSETS_DIR = BASE_DIR / "assets"

LOGO_PATH = ASSETS_DIR / "logo.png"
ICON_PATH = ASSETS_DIR / "page_icon.png"


# -------------------------------------------------------------------
# Page Setup
# -------------------------------------------------------------------
def init_page(page_title: str, layout: str = "wide", logo_link: str = "https://dxc.com/uk/en"):
    st.set_page_config(
        page_title=page_title,
        page_icon=Image.open(ICON_PATH),
        layout=layout
    )
    st.logo(
        Image.open(LOGO_PATH),
        link=logo_link,
        icon_image=Image.open(ICON_PATH),
        size="large"
    )
    st.title(page_title)

# -------------------------------------------------------------------
# Page Setup
# -------------------------------------------------------------------
def hide_sidebar():
    st.markdown("""
    <style>
        [data-testid="stSidebar"] {display: none;}
    </style>
""", unsafe_allow_html=True)

# -------------------------------------------------------------------
# Access Control
# -------------------------------------------------------------------
def require_access(user_key="user_emails", admin_key="admin_emails"):
    access = st.secrets.get("access_control", {})
    user_email = st.session_state.get("user_email", "").lower()
    user_name = st.session_state.get("user_name", "").strip()

    allowed_users = {e.lower() for e in access.get(user_key, [])}
    admins = {e.lower() for e in access.get(admin_key, [])}

    is_admin = user_email in admins
    is_authorised = user_email in allowed_users or is_admin

    if not is_authorised:
        st.error("You do not have permission to access this page.")
        st.stop()

    return user_email, user_name, is_admin


# -------------------------------------------------------------------
# Authenticated Supabase client
# -------------------------------------------------------------------
def get_authenticated_supabase():
    token = st.session_state.get("token")

    if not token or "id_token" not in token:
        st.error("Authentication expired — log in again.")
        st.stop()

    client = get_supabase()

    try:
        client.auth.sign_in_with_id_token({
            "provider": "azure",
            "id_token": token["id_token"],
            "token": token["access_token"]
        })
    except Exception as e:
        st.error("Supabase authentication failed — please re-login.")
        st.exception(e)
        st.stop()

    return client


# -------------------------------------------------------------------
# Sidebar User Info
# -------------------------------------------------------------------
def render_user_sidebar(user_name, user_email):
    with st.sidebar:
        st.success(f"Welcome {user_name}")
        st.caption(f"Signed in as: {user_email}")

        st.markdown("---")
        if st.button("Logout"):
            st.session_state.clear()
            st.rerun()


# -------------------------------------------------------------------
# Convert Member Rows → DataFrame for Data Editor
# -------------------------------------------------------------------
def members_to_dataframe(members, team_lookup=None):
    team_lookup = team_lookup or {}

    df = pd.DataFrame([{
        "id": str(m.get("id")),
        "Team Name": team_lookup.get(m.get("team_id"), "Unassigned"),
        "Role": m.get("role"),
        "Preferred Route": m.get("preferred_route"),
        "On Waiting List": bool(m.get("on_waiting_list")),
        "Full Name": m.get("full_name"),
        "Employee ID": m.get("employee_id"),
        "Employee Email": m.get("employee_email"),
        "Mobile Number": m.get("mobile_number"),
        "Organisation": m.get("organisation"),
        "Shirt Size": m.get("shirt_size"),

        # Booleans + Text
        "Forces Veteran": bool(m.get("forces_vet")),
        "Camping Friday": bool(m.get("camping_fri")),
        "Camping Saturday": bool(m.get("camping_sat")),
        "Taking Car": bool(m.get("taking_car")),
        "Notes": m.get("notes"),
        "Hiking Experience": m.get("hiking_experience"),
        "Travelling From": m.get("travelling_from")

    } for m in members])

    return df


# -------------------------------------------------------------------
# Excel Export Helpers (styling + formatting)
# -------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def _style_worksheet(ws):
    """Apply header styling, borders, and column sizing to worksheet."""
    # Style header row
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    # Apply borders to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(15, max_len + 2)


# -------------------------------------------------------------------
# Excel Export
# -------------------------------------------------------------------
def export_excel(client):
    teams = client.table("teams").select("id, team_name, route").execute().data or []
    members = client.table("members").select("*").execute().data or []
    team_lookup = {t["id"]: t["team_name"] for t in teams}

    wb = Workbook()
    ws = wb.active
    ws.title = "All Members"

    HEADERS = [
        "Team Name", "On Waiting List", "Name", "Email", "Mobile Number", "Preferred Route", "Organisation",
        "Role", "Shirt Size", "Forces Veteran", "Camping Friday",
        "Camping Saturday", "Taking Car", "Hiking Experience",
        "Travelling From", "Notes"
    ]

    ws.append(HEADERS)
    _style_worksheet(ws)

    # Sort members: by team name (Unassigned last), then by on_waiting_list (False first)
    sorted_members = sorted(
        members,
        key=lambda m: (
            team_lookup.get(m.get("team_id"), "Unassigned") == "Unassigned",  # Unassigned goes to end
            team_lookup.get(m.get("team_id"), "Unassigned"),  # Then sort by team name
            bool(m.get("on_waiting_list", False)),  # Then False (active) before True (waiting)
        )
    )

    for m in sorted_members:
        ws.append([
            team_lookup.get(m.get("team_id"), "Unassigned"),
            "Yes" if m.get("on_waiting_list") else "No",
            m.get("full_name", ""),
            m.get("employee_email", ""),
            m.get("mobile_number", ""),
            m.get("preferred_route", ""),
            m.get("organisation", ""),
            m.get("role", ""),
            m.get("shirt_size", ""),
            m.get("forces_vet", False),
            m.get("camping_fri", False),
            m.get("camping_sat", False),
            m.get("taking_car", False),
            m.get("hiking_experience", ""),
            m.get("travelling_from", ""),
            m.get("notes", "")
        ])

    _style_worksheet(ws)
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# -------------------------------------------------------------------
# Update Members (called after data_editor editing)
# -------------------------------------------------------------------
COLUMN_MAP = {
    "Full Name": "full_name",
    "Employee ID": "employee_id",
    "Employee Email": "employee_email",
    "Mobile Number": "mobile_number",
    "Organisation": "organisation",
    "Preferred Route": "preferred_route",
    "On Waiting List": "on_waiting_list",
    "Role": "role",
    "Shirt Size": "shirt_size",
    "Forces Veteran": "forces_vet",
    "Camping Friday": "camping_fri",
    "Camping Saturday": "camping_sat",
    "Taking Car": "taking_car",
    "Notes": "notes",
    "Hiking Experience": "hiking_experience",
    "Travelling From": "travelling_from",
}


def apply_member_updates(edited_df, original_df, team_name_to_id, client):
    if edited_df is None or edited_df.empty:
        return 0

    updates = 0

    # Convert IDs to strings once
    original_df["id"] = original_df["id"].astype(str)
    edited_df["id"] = edited_df["id"].astype(str)

    for _, row in edited_df.iterrows():
        member_id = row["id"]
        orig = original_df[original_df["id"] == member_id].iloc[0]

        changes = {}

        for col, new_val in row.items():
            if col == "id":
                continue

            old_val = orig[col]

            # TEAM NAME special case
            if col == "Team Name":
                mapped = None if new_val == "Unassigned" else team_name_to_id.get(new_val)
                if new_val != old_val:
                    changes["team_id"] = mapped
                continue

            # Regular field update
            db_col = COLUMN_MAP.get(col)
            if db_col and new_val != old_val:
                changes[db_col] = new_val

        if changes:
            try:
                client.table("members").update(changes).eq("id", member_id).execute()
                updates += 1
            except Exception as e:
                st.error(f"Failed to update {member_id}: {e}")

    return updates


def delete_member(member_id: int, client):
    try:
        client.table("members").delete().eq("id", member_id).execute()
        st.success("Member deleted.")
        st.rerun()
    except Exception as e:
        st.error(f"Failed to delete member: {e}")


def delete_team(team_id: int, client):
    try:
        # Unassign members instead of deleting them
        client.table("members").update({"team_id": None}).eq("team_id", team_id).execute()

        # Delete the actual team
        client.table("teams").delete().eq("id", team_id).execute()

        st.success("Team deleted.")
        st.rerun()
    except Exception as e:
        st.error(f"Failed to delete team: {e}")


# -------------------------------------------------------------------
# Helpers: prepare/sanitize member records for DB and counts
# -------------------------------------------------------------------
def get_active_member_count(client) -> int:
    """Return the count of members not on the waiting list."""
    try:
        res = client.table("members").select("id", count="exact").eq("on_waiting_list", False).execute()
        return res.count or 0
    except Exception:
        return 0


def prepare_member_record(draft: dict, on_waiting_list: bool | None = None, client=None) -> dict:
    """
    Creates a sanitized DB-ready member record from the session draft.

    Fixes:
    - Allows team_id=None to overwrite old DB values (important for switching to Independent)
    - Only includes allowed DB columns
    - Sanitizes text inputs
    - If team_id missing but team_name exists, resolves ID from DB
    """

    # Allowed DB columns
    allowed = {
        "full_name", "employee_email", "employee_id",
        "team_id", "preferred_route", "role",
        "shirt_size", "forces_vet", "camping_fri", "camping_sat",
        "taking_car", "travelling_from", "notes", "hiking_experience",
        "mobile_number", "organisation", "on_waiting_list"
    }

    record = {}

    # -------------------------------------------------------
    # If team_id missing but team_name exists → resolve ID
    # -------------------------------------------------------
    if "team_id" not in draft and draft.get("team_name") and client:
        try:
            t = (
                client.table("teams")
                .select("id")
                .eq("team_name", draft.get("team_name"))
                .execute()
            )
            if t.data:
                record["team_id"] = t.data[0]["id"]
        except Exception:
            pass

    # -------------------------------------------------------
    # Copy values from draft (including None!)
    # -------------------------------------------------------
    for k in allowed:
        if k in draft:
            val = draft[k]

            # Normalise booleans
            if k in {"forces_vet", "camping_fri", "camping_sat", "taking_car", "on_waiting_list"}:
                record[k] = bool(val)
                continue

            # Allow None for team_id (so Independent updates properly)
            if k == "team_id":
                record[k] = val  # can be None
                continue

            # Clean text values
            if isinstance(val, str):
                from helpers import sanitize_text
                record[k] = sanitize_text(val)
            else:
                record[k] = val

    # Explicit override of waiting flag
    if on_waiting_list is not None:
        record["on_waiting_list"] = bool(on_waiting_list)

    return record



# -------------------------------------------------------------------
# GPX Utilities: Parsing & Distance/Elevation Calculations
# -------------------------------------------------------------------
def load_gpx_points(path):
    """
    Load GPX points from a file path.
    Returns a list of tuples (lat, lon, ele). Supports <trkpt> and <rtept>.
    """
    try:
        if not os.path.exists(path):
            return None
        with open(path, "rb") as f:
            root = ET.parse(f).getroot()

        points = []

        # Track points (preferred)
        for trk in root.findall(".//{*}trk"):
            for seg in trk.findall(".//{*}trkseg"):
                for pt in seg.findall(".//{*}trkpt"):
                    lat = float(pt.get("lat"))
                    lon = float(pt.get("lon"))
                    ele_el = pt.find(".//{*}ele")
                    ele = float(ele_el.text) if ele_el is not None else None
                    points.append((lat, lon, ele))

        # Route points (fallback)
        if not points:
            for pt in root.findall(".//{*}rtept"):
                lat = float(pt.get("lat"))
                lon = float(pt.get("lon"))
                ele_el = pt.find(".//{*}ele")
                ele = float(ele_el.text) if ele_el is not None else None
                points.append((lat, lon, ele))

        return points if points else None
    except Exception:
        return None


def haversine_m(lat1, lon1, lat2, lon2):
    """Great-circle distance between two lat/lon points in meters."""
    R = 6371000.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlmb / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


def compute_track_stats(points):
    """
    Compute distance (km), ascent/descent (m), and cumulative arrays for plotting.
    points: [(lat, lon, ele), ...]
    """
    if not points or len(points) < 2:
        return {
            "distance_km": 0.0,
            "ascent_m": 0.0,
            "descent_m": 0.0,
            "cum_dist_km": [],
            "elev": [],
        }

    distance_m = 0.0
    ascent_m = 0.0
    descent_m = 0.0

    cum_dist_km = [0.0]
    elev = [points[0][2] if points[0][2] is not None else None]

    for i in range(1, len(points)):
        lat1, lon1, ele1 = points[i - 1]
        lat2, lon2, ele2 = points[i]

        d = haversine_m(lat1, lon1, lat2, lon2)
        distance_m += d
        cum_dist_km.append(distance_m / 1000.0)

        if ele1 is not None and ele2 is not None:
            delta = ele2 - ele1
            if delta > 0:
                ascent_m += delta
            elif delta < 0:
                descent_m += -delta
            elev.append(ele2)
        else:
            elev.append(None)

    return {
        "distance_km": distance_m / 1000.0,
        "ascent_m": ascent_m,
        "descent_m": descent_m,
        "cum_dist_km": cum_dist_km,
        "elev": elev,
    }


def trim_outliers(points, q=0.01):
    """
    Trim GPS outliers by removing points outside the [q, 1-q] quantile range
    in BOTH latitude and longitude. If trimming would leave < 2 points,
    fall back to original points.
    """
    if not points or len(points) < 10:
        return points

    lats = sorted(p[0] for p in points)
    lons = sorted(p[1] for p in points)
    n = len(points)
    low_idx = max(0, int(n * q))
    high_idx = min(n - 1, int(n * (1 - q)) - 1)

    low_lat, high_lat = lats[low_idx], lats[high_idx]
    low_lon, high_lon = lons[low_idx], lons[high_idx]

    trimmed = [p for p in points if (low_lat <= p[0] <= high_lat) and (low_lon <= p[1] <= high_lon)]
    return trimmed if len(trimmed) >= 2 else points


def expanded_bounds(points, margin_frac=0.08):
    """
    Compute min/max lat/lon and expand by margin_frac (default 8%) so
    the route fits comfortably inside the frame.
    """
    if not points:
        # Fallback bounds to avoid downstream errors
        return (54.46 - 0.001, -3.02 - 0.001, 54.46 + 0.001, -3.02 + 0.001)

    lats = [p[0] for p in points]
    lons = [p[1] for p in points]

    min_lat, max_lat = min(lats), max(lats)
    min_lon, max_lon = min(lons), max(lons)

    lat_span = max_lat - min_lat
    lon_span = max_lon - min_lon

    # If span is zero (e.g., identical points), use small epsilon to avoid identical bounds
    eps_lat = lat_span * margin_frac if lat_span > 0 else 0.001
    eps_lon = lon_span * margin_frac if lon_span > 0 else 0.001

    return (min_lat - eps_lat, min_lon - eps_lon, max_lat + eps_lat, max_lon + eps_lon)


def mid_route_center(points, cum_dist_km):
    """
    Return the lat/lon of the point closest to half the total route distance.
    """
    if not points or not cum_dist_km:
        return points[0][:2] if points else (54.46, -3.02)

    total = cum_dist_km[-1]
    target = total / 2.0
    idx = min(range(len(cum_dist_km)), key=lambda i: abs(cum_dist_km[i] - target))
    lat, lon, _ = points[idx]
    return (lat, lon)


# -------------------------------------------------------------------
# SIMPLE BACK BUTTON HELPER
# -------------------------------------------------------------------
def back_button(target_page):
    st.markdown("---")
    if st.button("← Previous Page"):
        st.switch_page(target_page)


# GENERIC SANITIZATION FUNCTION USED FOR ALL SANITIZATIONS OF TEXT INPUT
def sanitize_text(value: str) -> str:
    """Strip whitespace, normalize unicode, remove control chars."""
    if not value:
        return ""
    value = unicodedata.normalize("NFKC", value)
    value = re.sub(r"[\x00-\x1f\x7f]", "", value)
    return value.strip()